import sys
import argparse
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.keys import Keys
import time
import undetected_chromedriver as uc
import openpyxl
import re
import os
import pandas as pd
import random

def calculate_average(text):
    if isinstance(text, str):
        numbers = re.findall(r'[\d,]+(?:\.\d+)?', text)
        if len(numbers) == 2:
            num1 = int(numbers[0].replace(',', ''))
            num2 = int(numbers[1].replace(',', ''))
            return (num1 + num2) / 2
    return None


def get_property_type(title):
    if not title:
        return ""
    words = title.split()
    if (('أرض' in words or 'ارض' in words) and 'للبيع' in words):
        property_type = 'أرض للبيع'
    elif ('شقة' in words and 'للإيجار' in words):
        property_type = 'شقة للإيجار'
    elif ('شقة' in words and 'للبيع' in words):
        property_type = 'شقة للبيع'
    elif ('فيلا' in words and 'للبيع' in words):
        property_type = 'فيلا للبيع'
    elif ('فيلا' in words and 'للإيجار' in words):
        property_type = 'فيلا للإيجار'
    elif ('عمارة' in words and 'للبيع' in words):
        property_type = 'عمارة للبيع'
    elif ('دور' in words and 'للبيع' in words):
        property_type = 'دور للبيع'
    elif ('دور' in words and 'للإيجار' in words):
        property_type = 'دور للإيجار'
    elif ('غرفة' in words and 'للإيجار' in words):
        property_type = 'غرفة للإيجار'
    elif ('ستوديو' in words and 'للإيجار' in words):
        property_type = 'ستوديو للإيجار'
    elif ('محل' in words and 'للإيجار' in words):
        property_type = 'محل للإيجار'
    else:
        property_type = ""
    return property_type

# اسم الملف
excel_file = "dealapp_data.xlsx"

# التحقق من وجود الملف
if os.path.exists(excel_file):
    workbook = openpyxl.load_workbook(excel_file) # فتح ملف موجود
    sheet = workbook.active
    print("Excel file loaded successfully. Data will be appended.")
else:
    workbook = openpyxl.Workbook() # انشاء ملف جديد
    sheet = workbook.active
    sheet.append(["الرابط", "العنوان", "نوع العقار", "المدينة", "الأحياء", "السعر", "الغرض", "نوع العقد", "الغرف", "دورات المياه", "الصالات", "التفاصيل", "المساحة","عرض الشارع"])
    print("New Excel file created.")


chrome_options = uc.ChromeOptions()
# chrome_options.add_argument("--headless")
# chrome_options.add_argument("--disable-blink-features=AutomationControlled")
# chrome_options.add_argument("--disable-gpu")
# chrome_options.add_argument("--no-sandbox")
# chrome_options.add_argument("--disable-dev-shm-usage")
# chrome_options.add_argument("--disable-infobars")
# chrome_options.add_argument("--disable-extensions")
# chrome_options.add_argument("--disable-popup-blocking")
# chrome_options.add_argument("--disable-notifications")
chrome_options.add_argument("--enable-javascript")
# chrome_options.add_argument("--disable-blink-features=AutomationControlled")
# chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
driver = uc.Chrome(options=chrome_options)

num_pages = 5
existing_data = [] #لتخزين بيانات الجدول الموجوده بالفعل
#تخزين البيانات الموجوده بالفعل قبل الاضافة عليها
if sheet.max_row > 1:
    for row in sheet.iter_rows(min_row=2, values_only=True):
        existing_data.append(list(row))

for page in range(num_pages):
    print(f"Navigating to page {page + 1}")
    retries = 3  # عدد محاولات إعادة التحميل
    retry_count = 0
    page_loaded_successfully = False

    while retry_count < retries and not page_loaded_successfully:
        random_number = random.randint(100000, 999999)
        url = f"https://dealapp.sa/request-details/{random_number}"
        print(f"Trying URL: {url}, Retry attempt: {retry_count + 1}")
        try:
            driver.get(url)
            time.sleep(5) # وقت إضافي للتحميل والتأكد من ظهور العناصر

            # محاولة العثور على عنصر رئيسي في الصفحة للتأكد من تحميلها بشكل صحيح
            driver.find_element(By.ID, "request-title") # عنصر العنوان كمؤشر لنجاح التحميل

            link = driver.current_url
            title_element = driver.find_element(By.ID, "request-title")
            title = title_element.text.strip()
            time.sleep(2)
            row_data = []
            property_type = get_property_type(title)

            city_element = driver.find_element(By.ID, "request-city")
            city_name = city_element.text.strip()
            districts_element = driver.find_element(By.ID, "request-districts")
            districts = districts_element.text.strip()
            price_element = driver.find_element(By.ID, "price")
            price = price_element.text.strip()

            try:
                contract_type_element = driver.find_element(By.ID, "contractType")
                contract_type = contract_type_element.text.strip()
            except:
                contract_type = 0
            try:
                purpose_element = driver.find_element(By.ID, "purpose")
                purpose = purpose_element.text.strip()
            except:
                purpose = 0

            element = driver.find_element(By.ID, "price")
            driver.execute_script("arguments[0].scrollIntoView();", element)
            time.sleep(2)

            try:
                rooms_num_element = driver.find_element(By.ID, "roomsNum-value")
                rooms_num = rooms_num_element.text.strip()
                bathroom_element = driver.find_element(By.ID, "bathroom-value")
                bathroom = bathroom_element.text.strip()
                hall_element = driver.find_element(By.ID, "hallsNum-value")
                halls_num = hall_element.text.strip()
            except:
                rooms_num = bathroom = halls_num = 0

            try:
                area_element = driver.find_element(By.ID, "area-value")
                area = area_element.text.strip()
            except:
                area = 0
            try:
                street_element = driver.find_element(By.ID, "streetvalue-value")
                street = street_element.text.strip()
            except:
                street = 0

            details_element = driver.find_element(By.XPATH, '//div[@id="request-note-value"]//p')
            details = details_element.text.strip().replace('\n', ', ')

            row_data = [link, title, property_type, city_name, districts, price, purpose, contract_type, rooms_num, bathroom, halls_num, details, area, street]

            if row_data not in existing_data:
                sheet.append(row_data)
                existing_data.append(row_data)
            else :
                print("Duplicate Data")

            page_loaded_successfully = True # تم تحميل الصفحة بنجاح والخروج من حلقة المحاولة
            print(f"Page {page + 1} loaded successfully after {retry_count + 1} attempts.")


        except Exception as e:
            print(f"Error loading page with URL: {url}, error: {e}")
            retry_count += 1
            time.sleep(5) # انتظار قبل إعادة المحاولة

    if not page_loaded_successfully:
        print(f"Failed to load page {page + 1} after {retries} retries. Moving to the next page or ending.")


    # الانتقال للصفحة التالية (إذا لم نكن في الصفحة الأخيرة)
    if page < num_pages - 1 and page_loaded_successfully: # فقط ننتقل للصفحة التالية إذا تم تحميل الصفحة الحالية بنجاح
        wait = WebDriverWait(driver, 10)
        button = wait.until(EC.element_to_be_clickable((By.CLASS_NAME, 'btn')))
        button.click()
        time.sleep(5) # وقت للتحميل بعد النقر على زر الصفحة التالية


print(f"Navigated to {num_pages} pages (or attempted to) successfully.")



    # حفظ ملف Excel
workbook.save(excel_file)
driver.quit()
df = pd.read_excel(excel_file)
df = df.drop_duplicates()
df['السعر'] = df['السعر'].apply(calculate_average)
df['المساحة'] = df['المساحة'].apply(calculate_average)
df['عرض الشارع'] = df['عرض الشارع'].apply(calculate_average)
df = df.fillna(0)
df.to_excel(excel_file, index=False)

# _______________________________________________________________________________________________________________
# _______________________________________________________________________________________________________________
# _______________________________________________________________________________________________________________
# _______________________________________________________________________________________________________________


chrome_options = uc.ChromeOptions()
# chrome_options.add_argument("--headless")
# chrome_options.add_argument("--disable-blink-features=AutomationControlled")
# chrome_options.add_argument("--disable-gpu")
# chrome_options.add_argument("--no-sandbox")
# chrome_options.add_argument("--disable-dev-shm-usage")
# chrome_options.add_argument("--disable-infobars")
# chrome_options.add_argument("--disable-extensions")
# chrome_options.add_argument("--disable-popup-blocking")
# chrome_options.add_argument("--disable-notifications")
chrome_options.add_argument("--enable-javascript")
# chrome_options.add_argument("--disable-blink-features=AutomationControlled")
# chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
driver = uc.Chrome(options=chrome_options)

driver.get("https://sa.aqar.fm/login")
time.sleep(1)

# --- استقبال الإيميل والباسورد كـ arguments ---
parser = argparse.ArgumentParser(description='Scrape aqar.fm with email and password.')
parser.add_argument('--email', type=str, help='Email for aqar.fm login')
parser.add_argument('--password', type=str, help='Password for aqar.fm login')
args = parser.parse_args()

email_from_args = args.email
password_from_args = args.password

# --- استخدام الإيميل والباسورد من arguments أو القيم الافتراضية إذا لم يتم تمريرها ---
email_login = email_from_args if email_from_args else "966542266667" # قيمة افتراضية إذا لم يتم تمرير argument
password_login = password_from_args if password_from_args else "KareeM854120" # قيمة افتراضية إذا لم يتم تمرير argument


# 3️⃣ العثور على حقول الإدخال وإدخال البيانات
email_input = driver.find_element(By.NAME, "phone")
password_input = driver.find_element(By.NAME, "password")

email_input.send_keys(email_login) # استخدام الإيميل من arguments
password_input.send_keys(password_login) # استخدام الباسورد من arguments

# 4️⃣ الضغط على زر تسجيل الدخول
password_input.send_keys(Keys.RETURN)
time.sleep(2)

# 5️⃣ إنشاء/فتح ملف إكسل وكتابة العناوين
excel_file = "aqar_data.xlsx"
workbook = openpyxl.Workbook()
if os.path.exists(excel_file):
    workbook = openpyxl.load_workbook(excel_file)
sheet = workbook.active
if sheet.max_row == 1 :
   sheet.append(["الرابط", "العنوان", "السعر", "المساحة", "عرض الشارع", "الغرض","الغرف", "الصالات", "الحمامات", "الحي", "المدينة", "مدة الايجار", "نوع العقار"]) # كتابة عناوين الأعمدة

base_urls = [
    "https://sa.aqar.fm/%D8%B4%D9%82%D9%82-%D9%84%D9%84%D8%A5%D9%8A%D8%AC%D8%A7%D8%B1/",
    "https://sa.aqar.fm/%D8%B4%D9%82%D9%82-%D9%84%D9%84%D8%A8%D9%8A%D8%B9/",
    "https://sa.aqar.fm/%D9%81%D9%84%D9%84-%D9%84%D9%84%D8%A8%D9%8A%D8%B9/",
    "https://sa.aqar.fm/%D9%81%D9%84%D9%84-%D9%84%D9%84%D8%A5%D9%8A%D8%AC%D8%A7%D8%B1/",
    "https://sa.aqar.fm/%D8%A3%D8%B1%D8%A7%D8%B6%D9%8A-%D9%84%D9%84%D8%A8%D9%8A%D8%B9/",
    "https://sa.aqar.fm/%D8%AF%D9%88%D8%B1-%D9%84%D9%84%D8%A5%D9%8A%D8%AC%D8%A7%D8%B1/",
    "https://sa.aqar.fm/%D8%AF%D9%88%D8%B1-%D9%84%D9%84%D8%A8%D9%8A%D8%B9/",
    "https://sa.aqar.fm/%D8%BA%D8%B1%D9%81-%D9%84%D9%84%D8%A5%D9%8A%D8%AC%D8%A7%D8%B1/"
]

def extract_location(text):
    pattern = r"حي ([^,]+), مدينة ([^,]+)"
    match = re.search(pattern, text)
    if match:
        return match.group(1), match.group(2)
    else:
        return 0, 0

existing_links = set()
for row in sheet.iter_rows(min_row=2, values_only=True):
    if row and row[0]:  # تأكد من وجود بيانات في الصف وأن الرابط ليس فارغًا
        existing_links.add(row[0])


for base_url in base_urls:
    page_number = 1

    while page_number <= 2:

        driver.get(f"{base_url}{page_number}?unverified_owner=eq,1")
        time.sleep(3)

        listings_div = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, "_list__Ka30R"))
        )
        if listings_div.text == '':
            break

        listings = listings_div.find_elements(By.CSS_SELECTOR, "div > div > a")

        for listing in listings:
            link = listing.get_attribute("href")

            # Check if the link already exists in the Excel file
            if link in existing_links:
                # print(f"Skipping duplicate link: {link}")
                continue

            title = listing.find_element(By.CSS_SELECTOR, "h4").text
            price = listing.find_element(By.CLASS_NAME, "_price__X51mi").text
            specs_div = listing.find_element(By.CLASS_NAME, "_specs__nbsgm")
            spec_items = specs_div.find_elements(By.CLASS_NAME, "_spec__SIJiK")

            area = rooms = halls = bathrooms = street = type = 0
            for item in spec_items:
                img = item.find_element(By.TAG_NAME, "img")
                icon_src = img.get_attribute("src")
                value = item.text.strip()

                if "area.svg" in icon_src:
                    area = value
                elif "bed-king.svg" in icon_src:
                    rooms = value
                elif "couch.svg" in icon_src:
                    halls = value
                elif "bath.svg" in icon_src:
                    bathrooms = value
                elif "street.svg" in icon_src:
                    street = value
                elif "pinned-note.svg" in icon_src:
                    type = value

            neighborhood, city = extract_location(title)
            rent_duration = ""
            if "/شهري" in price:
                rent_duration = "شهري"
                price = price.replace('/شهري', '')
            elif "/سنوي" in price:
                rent_duration = "سنوي"
                price = price.replace('/سنوي', '')

            price = str(price).replace('ريال', '').replace(',', '')
            try:
                price = float(price)
            except ValueError:
                price = None

            area, street = str(area), str(street)
            area = re.search(r'(\d+)', area)
            street = re.search(r'(\d+)', street)
            if area:
                area = int(area.group(1))
            else:
                area = None

            if  street:
                street = int(street.group(1))
            else:
                street = None

            rooms = int(rooms) if rooms else 0
            halls = int(halls) if halls else 0
            bathrooms = int(bathrooms) if bathrooms else 0

            property_type = " ".join(title.split()[:2]) if title else ""
            new_row = [link, title, price, area, street, type, rooms, halls, bathrooms, neighborhood, city, rent_duration, property_type]
            sheet.append(new_row)
            existing_links.add(link)
        page_number +=1



workbook.save(excel_file)

driver.quit()

df = pd.read_excel(excel_file)

def replace_ha(city):
  if city != 0:
    if city == 'مكه المكرمه':
        return 'مكة المكرمة'
    elif city == 'المدينه المنوره':
       return 'المدينة المنورة'
    else:
        return city


df['المدينة'] = df['المدينة'].apply(replace_ha)
df.to_excel(excel_file, index=False)
