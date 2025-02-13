from flask import Flask, render_template, request, jsonify, session, redirect, url_for
import pandas as pd
pd.set_option('future.no_silent_downcasting', True)
import openpyxl
import os
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import re
import threading
import time
from selenium.webdriver.chrome.options import Options
import numpy as np
import subprocess
import configparser 

app = Flask(__name__)
app.secret_key = 'your_secret_key'

# مسارات الملفات
order_df_path = 'orders.xlsx'
aqar_data_path = 'aqar_data.xlsx'
order2_df_path = 'dealapp_data.xlsx'

PRICE_THRESHOLD = 0.4
SIZE_THRESHOLD = 50
ROOM_THRESHOLD = 0
BATHROOM_THRESHOLD = 0
HALL_THRESHOLD = 0
STREET_THRESHOLD = 4

# -- الدوال المساعدة (load_orders2_data, get_order_data, load_data, load_excel_data, find_similar_listings, find_similar_for_row) --
# ... (نفس الدوال المساعدة اللي كتبناها قبل كده) ...
def load_orders2_data():
    """تحميل البيانات من ملف Excel الثاني مع معالجة الأخطاء."""
    try:
        order2_data = pd.read_excel(order2_df_path)
        return order2_data
    except FileNotFoundError as e:
        print(f"Error loading files: {e}")
        return None

def get_order_data(source, row_index):
        """البحث عن عقارات مشابهة لصف معين."""
        try:
            if source == 'order1':
                df, order_data = load_data()
            elif source == 'order2':
                 df = pd.read_excel('aqar_data.xlsx')
                 order_data = load_orders2_data()
            else:
                raise ValueError("مصدر البيانات غير صالح")
            if order_data is None or df is None:
                raise ValueError("خطأ في تحميل البيانات.")

            if row_index < 0 or row_index >= len(order_data):
                raise ValueError("رقم الصف خارج الحدود.")
            request_data = order_data.iloc[row_index].to_dict()

            similar_listings = find_similar_listings(df, request_data)
            # إنشاء نسخة من DataFrame قبل الحذف
            similar_listings_copy = similar_listings.copy()
            similar_listings_copy = similar_listings_copy.replace(0, np.nan)

            # حذف الأعمدة التي تحتوي على NaN
            similar_listings_copy = similar_listings_copy.dropna(axis=1, how='any')
            # تحويل البيانات المتشابهة إلى تنسيق JSON
            similar_listings_data = similar_listings_copy.to_dict(orient='records')
            return similar_listings_data
        except Exception as e:
            print(f"error:{e}")
            return None



def load_data():
    """تحميل البيانات من ملفات Excel مع معالجة الأخطاء."""
    try:
        df = pd.read_excel(aqar_data_path)
        order_data = pd.read_excel(order_df_path)
        return df, order_data
    except FileNotFoundError as e:
        print(f"Error loading files: {e}")
        return None, None


def load_excel_data(file):
    """تحميل البيانات من ملف Excel."""
    try:
        df = pd.read_excel(file)
        data = df.to_dict(orient='records')
        columns = df.columns.tolist()
        return data, columns
    except Exception as e:
        print(f"Error loading file: {e}")
        return None, None


def find_similar_listings(df, request):
    """البحث عن عقارات مشابهة."""
    # التأكد من أن البيانات من النوع الصحيح
    try:
        price = float(request['السعر'])
        size = int(request['المساحة'])
        rooms = int(request['الغرف'])
        bathrooms = int(request['دورات المياه'])
        halls = int(request['الصالات'])
        # property_type = request['نوع العقار']
        # street = request['عرض الشارع']
        city_r = request['المدينة']
    except (ValueError, TypeError):
        return pd.DataFrame()

    # استخراج أسماء الأعمدة
    price_col = 'السعر' if 'السعر' in df.columns else 'Price'
    size_col = 'المساحة' if 'المساحة' in df.columns else 'Size'
    room_col = 'الغرف' if 'الغرف' in df.columns else 'Rooms'
    bathroom_col = 'الحمامات' if 'الحمامات' in df.columns else 'Bathrooms'
    hall_col = 'الصالات' if 'الصالات' in df.columns else 'Halls'
    city_col = 'المدينة' if 'المدينة' in df.columns else 'City'
    # neighborhood_col = 'الحي' if 'الحي' in df.columns else 'Neighborhood'
    property_type_col = 'نوع العقار'
    street_col = 'عرض الشارع'

    if price != 0 :
        price_cond = (
            (df[price_col] >= price * (1 - PRICE_THRESHOLD)) &
            (df[price_col] <= price * (1 + PRICE_THRESHOLD))
        )
    else:
        price_cond = True

    if size != 0:
        size_cond = (
            (df[size_col] >= size - SIZE_THRESHOLD) &
            (df[size_col] <= size + SIZE_THRESHOLD)
        )
    else:
        size_cond = True

    if rooms != 0:
        room_cond = (
            (df[room_col] >= rooms - ROOM_THRESHOLD) &
            (df[room_col] <= rooms + ROOM_THRESHOLD)
        )
    else:
        room_cond = True

    if bathrooms != 0:
        bathroom_cond = (
            (df[bathroom_col] >= bathrooms - BATHROOM_THRESHOLD) &
            (df[bathroom_col] <= bathrooms + BATHROOM_THRESHOLD)
        )
    else:
        bathroom_cond = True

    if halls != 0:
        hall_cond = (
            (df[hall_col] >= halls - HALL_THRESHOLD) &
            (df[hall_col] <= halls + HALL_THRESHOLD)
        )
    else:
        hall_cond = True

    if 'عرض الشارع' in request:
        if request['عرض الشارع'] != 0:
            street_cond = (
                (df[street_col] >= request['عرض الشارع'] - STREET_THRESHOLD) &
                (df[street_col] <= request['عرض الشارع'] + STREET_THRESHOLD)
            )
        else:
            street_cond = True
    else:
        street_cond = True

    city = df[city_col] == city_r
    # neighborhood = df[neighborhood_col] == request[neighborhood_col]
    property = df['نوع العقار'] == request['نوع العقار']
    combined_cond = (
       property&
        size_cond &
        # price_cond &
        room_cond &
        bathroom_cond &
        hall_cond &
        city&
        street_cond
    )
    return df[combined_cond]


def find_similar_for_row(df, order_data, row_index):
    """البحث عن عقارات مشابهة لصف معين."""
    if row_index < 0 or row_index >= len(order_data):
        raise ValueError("رقم الصف خارج الحدود.")

    request_data = order_data.iloc[row_index].to_dict()

    similar_listings = find_similar_listings(df, request_data)

        # إنشاء نسخة من DataFrame قبل الحذف
    # similar_listings_copy = similar_listings.copy()
    # استبدال القيم الصفرية بـ NaN
    # similar_listings_copy = similar_listings_copy.replace(0, np.nan)
    # حذف الأعمدة التي تحتوي على NaN
    # similar_listings_copy = similar_listings_copy.dropna(axis=1, how='any')
    # تحويل البيانات المتشابهة إلى تنسيق JSON
    similar_listings_data = similar_listings.to_dict(orient='records')
    return similar_listings_data


# -- مسارات Flask --
@app.route('/')
def index():
    return render_template('index.html', current_page='index')


@app.route('/client')
def client():
    return render_template('client.html', current_page='client')


# إنشاء ملف Excel إذا لم يكن موجودًا
if not os.path.exists(order_df_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Real Estate Data"
    # إنشاء رؤوس الأعمدة
    sheet.append(["نوع العقار", "الاسم", "الجوال", "المساحة", "الصالات", "دورات المياه", "الغرف", "المدينة", "الحي", "السعر", "الحالة"])
    workbook.save(order_df_path)


@app.route('/submit', methods=['POST'])
def submit():
    # استلام البيانات من الفورم
    property_type = request.form.get('property_type')
    name = request.form.get('name')
    phone = request.form.get('phone')
    city = request.form.get('city')
    district = request.form.get('district')
    room = request.form.get('room', type=int) if request.form.get('room') else 0
    area = request.form.get('area', type=float)  if request.form.get('area') else 0.0
    Toilets = request.form.get('Toilets', type=int)  if request.form.get('Toilets') else 0
    hall = request.form.get('hall', type=int)  if request.form.get('hall') else 0
    price = request.form.get('price', type=float)

    # فتح ملف Excel وإضافة البيانات
    workbook = openpyxl.load_workbook(order_df_path)
    sheet = workbook.active
    sheet.append([property_type, name, phone, area, hall, Toilets, room, city, district, price, 'PENDING'])  # إضافة عمود الحالة
    workbook.save(order_df_path)

    return f"تم حفظ البيانات"


@app.route('/dashboard')
def dashboard():
    try:
        df, order_data = load_data()
        if order_data is None or df is None:
            return "Error: Could not load data."

        # حساب عدد الطلبات الكلية
        total_orders = len(order_data)

        # حساب عدد الطلبات المعلقة
        pending_orders = len(order_data[order_data['الحالة'] == 'PENDING'])

        # ترتيب البيانات بحيث تظهر حالات "DONE" أولاً
        order_data['الحالة'] = pd.Categorical(order_data['الحالة'], categories=['DONE', 'PENDING'], ordered=True)
        order_data = order_data.sort_values('الحالة')

        # تحويل البيانات إلى قائمة قواميس لسهولة العرض في القالب
        data = order_data.to_dict(orient='records')

        # الحصول على أسماء الأعمدة
        columns = order_data.columns.tolist()

        # تحميل البيانات من ملف Excel الثاني
        order2_data = load_orders2_data()
        order2_data_list = None
        order2_columns = None
        if order2_data is not None:
                order2_data_list = order2_data.to_dict(orient='records')
                order2_columns = order2_data.columns.tolist()

        return render_template('dashboard.html', data=data, columns=columns,
                                total_orders=total_orders, pending_orders=pending_orders, current_page='dashboard',
                                order2_data=order2_data_list, order2_columns=order2_columns
                                )
    except FileNotFoundError:
        return "Error: Excel file not found."
    except Exception as e:
        return f"An error occurred: {e}"


@app.route('/toggle_status', methods=['POST'])
def toggle_status():
    try:
        row_index = request.get_json()
        print(f"row_index: {row_index}")
        workbook = openpyxl.load_workbook(order_df_path)
        sheet = workbook.active
        print(f"sheet: {sheet}")

        # استخدم هذه الطريقة لحساب عدد الأعمدة
        num_cols = len(sheet[1])

        current_status = sheet.cell(row=row_index + 2, column=num_cols).value
        print(f"current_status: {current_status}")
        new_status = "DONE" if current_status == "PENDING" else "PENDING"
        sheet.cell(row=row_index + 2, column=num_cols).value = new_status
        workbook.save(order_df_path)
        print(f"new_status: {new_status}")
        return jsonify({"status": new_status})
    except Exception as e:
        print(f"error:{e}")
        return jsonify({"error": str(e)}), 500


@app.route('/process_row', methods=['POST'])
def process_row():
    try:
        row_index = request.get_json()
        source = request.args.get('source', 'order1')
        similar_data = get_order_data(source, row_index)
        if similar_data is None:
            return jsonify({"error": "Could not process data"}), 500
        return jsonify(similar_data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/delete_row', methods=['POST'])
def delete_row():
    try:
        row_index = request.get_json()
        workbook = openpyxl.load_workbook(order_df_path)
        sheet = workbook.active
        sheet.delete_rows(row_index + 2)  # يتم إضافة 2 لان الصفوف تبدأ من 1 و الصف الاول هو رؤوس الجدول.
        workbook.save(order_df_path)
        return jsonify({"success": True})
    except Exception as e:
        print(f"error:{e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/excel_data', methods=['GET', 'POST'])
def excel_data():
    if request.method == 'POST':
        if 'file' not in request.files:
            return render_template('excel_data.html', current_page='excel_data')
        file = request.files['file']
        if file.filename == '':
            return render_template('excel_data.html', current_page='excel_data')
        if file and file.filename.endswith(('.xlsx', '.xls')):
            try:
                data, columns = load_excel_data(file)
                if data and columns:
                    return render_template('excel_data.html', data=data, columns=columns, current_page='excel_data')
                else:
                    return "Error processing file"
            except Exception as e:
                return f"Error processing file: {e}"

    return render_template('excel_data.html', current_page='excel_data')

# --- دالة لقراءة بيانات الاعتماد من config.ini ---
def get_credentials():
    config = configparser.ConfigParser()
    config.read('config.ini') # اسم ملف الاعدادات
    if 'Credentials' in config:
        email = config['Credentials'].get('email', '') # القيمة الافتراضية في حال عدم وجود المتغير
        password = config['Credentials'].get('password', '') # القيمة الافتراضية في حال عدم وجود المتغير
        return email, password
    return '', '' # إرجاع قيم فارغة في حالة عدم وجود قسم 'Credentials'

# --- دالة لحفظ بيانات الاعتماد في config.ini ---
def save_credentials(email, password):
    config = configparser.ConfigParser()
    config['Credentials'] = {'email': email, 'password': password} # إنشاء القسم والمتغيرات
    with open('config.ini', 'w') as configfile: # فتح الملف للكتابة
        config.write(configfile) # كتابة التغييرات في الملف

    # --- تعديل مسار run_scraper ---
@app.route('/run_scraper', methods=['GET', 'POST'])
def run_scraper():
    output_message = ""
    email, password = get_credentials() # قراءة بيانات الاعتماد عند تحميل الصفحة

    if request.method == 'POST': # تشغيل السكربت فقط في حالة POST request (ضغط الزر)
        email = request.form.get('email')
        password = request.form.get('password')
        save_credentials(email, password) # حفظ بيانات الاعتماد الجديدة

        try:
            process = subprocess.Popen(['python', 'scraping.py', '--email', email, '--password', password],
                                        stdout=subprocess.PIPE,
                                        stderr=subprocess.PIPE,
                                        text=True)
            stdout, stderr = process.communicate()

            if process.returncode == 0:
                output_message = "تم تشغيل السكربت بنجاح!"
            else:
                output_message = "حدث خطأ أثناء تشغيل السكربت او ان اشتراكك في عقار انتهي"

        except Exception as e:
            output_message = "حدث خطأ غير متوقع او ان اشتراكك في عقار انتهي"

        # إعادة توجيه إلى نفس المسار بـ GET request بعد الـ POST
        return redirect(url_for('run_scraper', output_message=output_message, email=email, password=password))

    # في حالة الـ GET request (أو بعد الـ redirect)، نعرض الصفحة مع القيم الحالية
    output_message_from_redirect = request.args.get('output_message', '') # استخراج رسالة الحالة من query parameters إذا كانت موجودة
    email_from_redirect = request.args.get('email', email) # استخراج الإيميل من query parameters أو استخدام القيمة الحالية
    password_from_redirect = request.args.get('password', password) # استخراج الباسورد من query parameters أو استخدام القيمة الحالية

    return render_template('run_scraper.html',
                            current_page='run_scraper',
                            output_message=output_message_from_redirect or output_message, # استخدام رسالة الحالة من الـ redirect إذا كانت موجودة، أو الرسالة الحالية
                            email=email_from_redirect, # استخدام الإيميل من الـ redirect أو القيمة الحالية
                            password=password_from_redirect # استخدام الباسورد من الـ redirect أو القيمة الحالية
                            )

if __name__ == '__main__':
    app.run(debug=True)